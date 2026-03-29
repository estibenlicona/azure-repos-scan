"""Casos de uso de la aplicación."""

from __future__ import annotations

from collections import defaultdict
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Sequence, final

from azure_repos_scan.domain.models import (
    CodeSearchHit,
    DailyQueryRecord,
    DotNetVersion,
    MonthlySnapshot,
    OrganizationName,
    Project,
    RepoSummary,
    Repository,
    ScanResult,
)
from azure_repos_scan.domain.ports import AzureDevOpsClient, QueryStoragePort


@final
class ListProjectsUseCase:
    """Caso de uso: listar proyectos de la organización."""

    def __init__(self, client: AzureDevOpsClient) -> None:
        self._client = client

    def execute(self) -> Sequence[Project]:
        """Obtener todos los proyectos."""
        return self._client.list_projects()


@final
class ListRepositoriesUseCase:
    """Caso de uso: listar repositorios de un proyecto."""

    def __init__(self, client: AzureDevOpsClient) -> None:
        self._client = client

    def execute(self, project_name: str) -> Sequence[Repository]:
        """Obtener todos los repositorios de un proyecto."""
        return self._client.list_repositories(project_name)


@final
class ScanOrganizationUseCase:
    """Caso de uso: escanear todos los proyectos y repositorios de una organización."""

    def __init__(self, client: AzureDevOpsClient, organization: str) -> None:
        self._client = client
        self._organization = organization

    def execute(self) -> ScanResult:
        """Ejecutar escaneo completo de la organización."""
        projects = list(self._client.list_projects())
        all_repos: list[Repository] = []

        for project in projects:
            repos = self._client.list_repositories(project.name)
            all_repos.extend(repos)

        return ScanResult(
            organization=OrganizationName(self._organization),
            projects=projects,
            repositories=all_repos,
            scanned_at=datetime.now(timezone.utc),
            total_repos=len(all_repos),
            total_projects=len(projects),
        )


# ---------------------------------------------------------------------------
# Code Search .NET
# ---------------------------------------------------------------------------

ProgressCallback = Callable[[int, int, str], None]
"""callback(current_step, total_steps, message)"""


@final
class SearchDotNetProjectsUseCase:
    """Caso de uso: buscar proyectos .NET por versión usando Code Search."""

    def __init__(
        self,
        client: AzureDevOpsClient,
        store: QueryStoragePort,
        organization: str,
    ) -> None:
        self._client = client
        self._store = store
        self._organization = organization

    def execute(
        self,
        versions: Sequence[DotNetVersion],
        project: str | None,
        on_progress: ProgressCallback | None = None,
        branches: Sequence[str] = (),
    ) -> DailyQueryRecord:
        """Ejecutar búsqueda para las versiones seleccionadas.

        ``branches`` indica en cuáles ramas buscar; si está vacío se busca
        en todas las ramas indexadas sin filtro de branch.
        """
        results_by_version: dict[str, list[CodeSearchHit]] = {}
        queries = self._build_queries(versions)
        total_steps = len(queries)
        branch_list = list(branches)

        for step, (version, query) in enumerate(queries):
            if on_progress:
                on_progress(step, total_steps, f"Buscando {version.label}...")

            hits = self._fetch_all_pages(query, project, version, branch_list)
            key = version.moniker
            if key in results_by_version:
                existing = {
                    (h.repository_name, h.branch)
                    for h in results_by_version[key]
                }
                for h in hits:
                    pair = (h.repository_name, h.branch)
                    if pair not in existing:
                        results_by_version[key].append(h)
                        existing.add(pair)
            else:
                results_by_version[key] = hits

        if on_progress:
            on_progress(total_steps, total_steps, "Búsqueda completada")

        record = DailyQueryRecord(
            query_date=date.today(),
            organization=self._organization,
            project=project,
            versions_searched=list(versions),
            results_by_version=results_by_version,
        )
        self._store.save_query(record)
        return record

    @staticmethod
    def _build_queries(
        versions: Sequence[DotNetVersion],
    ) -> list[tuple[DotNetVersion, str]]:
        """Construir queries de búsqueda (singular + plural) por cada versión."""
        queries: list[tuple[DotNetVersion, str]] = []
        for v in versions:
            for q in v.build_search_queries():
                queries.append((v, q))
        return queries

    def _fetch_all_pages(
        self,
        query: str,
        project: str | None,
        version: DotNetVersion,
        branches: list[str],
    ) -> list[CodeSearchHit]:
        """Obtener repositorios que coinciden con la query usando facetas.

        Estrategia basada en facets (sin paginación):
        1. Llamada a nivel org con includeFacets → Project facets
        2. Por cada proyecto, llamada con Project en filtros → Repository facets
        3. Cada repo en las facetas genera un CodeSearchHit

        Si ``branches`` está vacío se consulta sin filtro de branch.
        Cuando hay branches se consulta cada una por separado para saber
        en cuál aparece cada repo.
        """
        if project:
            return self._fetch_for_project(query, project, version, branches)

        # Probe a nivel org para obtener proyectos vía facets
        base_filters: dict[str, list[str]] = {}
        if branches:
            base_filters["Branch"] = branches
        probe = self._client.search_code(
            query, None, skip=0, top=1,
            filters=base_filters or None, include_facets=True,
        )

        if probe.total_count == 0:
            return []

        # Obtener lista de proyectos desde facets
        project_names: list[str] = []
        if probe.facets and probe.facets.projects:
            project_names = [f.name for f in probe.facets.projects if f.result_count > 0]

        if not project_names:
            return []

        # Por cada proyecto, obtener repos
        all_hits: list[CodeSearchHit] = []
        for proj_name in project_names:
            all_hits.extend(
                self._fetch_for_project(query, proj_name, version, branches),
            )
        return all_hits

    def _fetch_for_project(
        self,
        query: str,
        project: str,
        version: DotNetVersion,
        branches: list[str],
    ) -> list[CodeSearchHit]:
        """Obtener hits de un proyecto, segmentando por branch si aplica."""
        if not branches:
            # Sin filtro de branch: una sola llamada
            filters = {"Project": [project]}
            probe = self._client.search_code(
                query, None, skip=0, top=1,
                filters=filters, include_facets=True,
            )
            return self._hits_from_repo_facets(probe, project, version, "")

        # Con branches: una llamada por branch para identificarlas
        all_hits: list[CodeSearchHit] = []
        for branch in branches:
            filters = {"Branch": [branch], "Project": [project]}
            probe = self._client.search_code(
                query, None, skip=0, top=1,
                filters=filters, include_facets=True,
            )
            all_hits.extend(
                self._hits_from_repo_facets(probe, project, version, branch),
            )
        return all_hits

    @staticmethod
    def _hits_from_repo_facets(
        page: CodeSearchPage,
        project_name: str,
        version: DotNetVersion,
        branch: str,
    ) -> list[CodeSearchHit]:
        """Construir hits desde las facetas de Repository."""
        if not page.facets or not page.facets.repositories:
            return []
        return [
            CodeSearchHit(
                repository_name=repo.name,
                project_name=project_name,
                dotnet_version=version,
                branch=branch,
            )
            for repo in page.facets.repositories
            if repo.result_count > 0
        ]


@final
class ExportResultsUseCase:
    """Caso de uso: exportar resultados a Excel."""

    def execute(self, hits: Sequence[CodeSearchHit], output_path: Path) -> Path:
        """Generar archivo Excel con los resultados."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = "Resultados .NET"

        # Encabezados
        headers = ["Repositorio", "Proyecto", "Versión .NET", "Branch"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Datos
        for row_idx, hit in enumerate(hits, start=2):
            ws.cell(row=row_idx, column=1, value=hit.repository_name)
            ws.cell(row=row_idx, column=2, value=hit.project_name)
            ws.cell(row=row_idx, column=3, value=hit.dotnet_version.label)
            ws.cell(row=row_idx, column=4, value=hit.branch)

        # Auto-ajustar anchos
        for col_idx in range(1, len(headers) + 1):
            letter = get_column_letter(col_idx)
            max_length = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[letter].width = min(max_length + 4, 60)

        wb.save(str(output_path))
        return output_path


# ---------------------------------------------------------------------------
# Dashboard / Reportes
# ---------------------------------------------------------------------------


@final
@dataclass(slots=True)
class DashboardData:
    """Datos procesados para el dashboard de reportes."""

    query_date: date
    organization: str
    total_repos: int
    total_csprojs: int
    repos_by_version: dict[DotNetVersion, list[RepoSummary]] = field(default_factory=dict)
    branches_available: list[str] = field(default_factory=list)
    branch_filter: str | None = None


def _group_hits_into_repos(
    hits: Sequence[CodeSearchHit],
    branch_filter: str | None = None,
) -> list[RepoSummary]:
    """Agrupar hits por repositorio, determinando la versión más antigua."""
    if branch_filter:
        hits = [h for h in hits if h.branch == branch_filter]

    groups: dict[tuple[str, str], list[CodeSearchHit]] = defaultdict(list)
    for hit in hits:
        groups[(hit.repository_name, hit.project_name)].append(hit)

    summaries: list[RepoSummary] = []
    for (repo_name, proj_name), repo_hits in groups.items():
        versions = frozenset(h.dotnet_version for h in repo_hits)
        branches = frozenset(h.branch for h in repo_hits)
        oldest = min(versions, key=lambda v: v.sort_key)
        summaries.append(
            RepoSummary(
                repository_name=repo_name,
                project_name=proj_name,
                oldest_version=oldest,
                all_versions=versions,
                branches=branches,
                csproj_count=len(repo_hits),
                hits=tuple(repo_hits),
            )
        )
    return summaries


@final
class BuildDashboardDataUseCase:
    """Caso de uso: construir datos del dashboard a partir de consultas almacenadas."""

    def __init__(self, store: QueryStoragePort) -> None:
        self._store = store

    def build_for_date(
        self,
        query_date: date,
        branch_filter: str | None = None,
    ) -> DashboardData | None:
        """Construir datos del dashboard para una fecha específica."""
        record = self._store.get_query_by_date(query_date)
        if record is None:
            return None

        all_hits = record.get_all_hits()
        summaries = _group_hits_into_repos(all_hits, branch_filter)

        repos_by_version: dict[DotNetVersion, list[RepoSummary]] = defaultdict(list)
        for s in summaries:
            repos_by_version[s.oldest_version].append(s)

        branches = sorted({h.branch for h in all_hits})

        return DashboardData(
            query_date=record.query_date,
            organization=record.organization,
            total_repos=len(summaries),
            total_csprojs=sum(s.csproj_count for s in summaries),
            repos_by_version=dict(repos_by_version),
            branches_available=branches,
            branch_filter=branch_filter,
        )

    def build_monthly_evolution(self, months: int = 6) -> list[MonthlySnapshot]:
        """Construir evolución mensual de repos por versión.

        Toma el record más reciente de cada mes.
        """
        records = self._store.load_queries()
        if not records:
            return []

        # Agrupar por mes, tomar el más reciente de cada mes
        by_month: dict[str, DailyQueryRecord] = {}
        for rec in records:
            month_key = rec.query_date.strftime("%Y-%m")
            if month_key not in by_month or rec.query_date > by_month[month_key].query_date:
                by_month[month_key] = rec

        # Ordenar meses y limitar
        sorted_months = sorted(by_month.keys())[-months:]

        snapshots: list[MonthlySnapshot] = []
        for month_key in sorted_months:
            rec = by_month[month_key]
            all_hits = rec.get_all_hits()
            summaries = _group_hits_into_repos(all_hits)
            version_counts: dict[DotNetVersion, int] = defaultdict(int)
            for s in summaries:
                version_counts[s.oldest_version] += 1
            snapshots.append(MonthlySnapshot(month=month_key, repos_by_version=dict(version_counts)))

        return snapshots

    def get_available_dates(self) -> list[date]:
        """Obtener las fechas disponibles en el historial, más reciente primero."""
        records = self._store.load_queries()
        dates = sorted({r.query_date for r in records}, reverse=True)
        return dates


@final
class ExportDashboardUseCase:
    """Caso de uso: exportar dashboard como imagen o PDF."""

    @staticmethod
    def to_image(widget: object, output_path: Path) -> Path:
        """Exportar un widget Qt como imagen PNG."""
        from PySide6.QtWidgets import QWidget

        if not isinstance(widget, QWidget):
            msg = "Se esperaba un QWidget"
            raise TypeError(msg)
        pixmap = widget.grab()
        pixmap.save(str(output_path), "PNG")
        return output_path

    @staticmethod
    def to_pdf(widget: object, output_path: Path) -> Path:
        """Exportar un widget Qt como PDF."""
        from PySide6.QtCore import QMarginsF
        from PySide6.QtGui import QPainter
        from PySide6.QtPrintSupport import QPrinter
        from PySide6.QtWidgets import QWidget

        if not isinstance(widget, QWidget):
            msg = "Se esperaba un QWidget"
            raise TypeError(msg)

        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(str(output_path))
        printer.setPageMargins(QMarginsF(10, 10, 10, 10))

        painter = QPainter()
        if not painter.begin(printer):
            msg = "No se pudo iniciar el painter"
            raise RuntimeError(msg)
        try:
            pixmap = widget.grab()
            page_rect = printer.pageRect(QPrinter.Unit.DevicePixel)
            scaled = pixmap.scaled(
                int(page_rect.width()),
                int(page_rect.height()),
                aspectMode=1,  # KeepAspectRatio
            )
            painter.drawPixmap(0, 0, scaled)
        finally:
            painter.end()
        return output_path
